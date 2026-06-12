"""
Sleeper Dynasty Fantasy League Tracker
Exports team rosters, draft picks, and free agents to Excel.

Usage:
    1. Set LEAGUE_ID below to your Sleeper league ID
       (found in URL: sleeper.com/leagues/<LEAGUE_ID>)
    2. Run: python3 sleeper_dynasty_tracker.py
    3. Opens dynasty_tracker.xlsx in the current directory
"""

import json
import os
import sys
import requests
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ─────────────────────────────────────────────
# CONFIG
# ─────────────────────────────────────────────
LEAGUE_ID     = "1322995024962543616"
SPORT         = "nfl"
OUTPUT        = "dynasty_tracker.xlsx"
PLAYERS_CACHE = "sleeper_players_cache.json"
KTC_CACHE     = "fantasycalc_cache.json"
# How many future rookie draft seasons to show (current year + N more)
PICK_SEASONS  = 3

# ── Dynamic stats season ──────────────────────────────────────────────────────
# NFL regular season runs September–January.
# • Sept–Dec  → current year's season is in progress  (show current year)
# • Jan–Aug   → offseason / last season is complete   (show previous year)
# This means STATS_SEASON and its cache auto-advance every September 1st
# with zero code changes required year over year.
def _resolve_stats_season() -> str:
    _now = datetime.now()
    return str(_now.year) if _now.month >= 9 else str(_now.year - 1)

STATS_SEASON = _resolve_stats_season()
STATS_CACHE  = f"sleeper_stats_cache_{STATS_SEASON}.json"
# ─────────────────────────────────────────────

BASE = "https://api.sleeper.app/v1"

# ── Styles ────────────────────────────────────────────────────────────────────

HEADER_FILL = PatternFill("solid", start_color="1F3864")
WHITE_FILL  = PatternFill("solid", start_color="FFFFFF")
HEADER_FONT = Font(name="Arial", bold=True, color="FFFFFF", size=10)
BODY_FONT   = Font(name="Arial", size=10)
BOLD_FONT   = Font(name="Arial", bold=True, size=10)
CENTER      = Alignment(horizontal="center", vertical="center")
LEFT        = Alignment(horizontal="left",   vertical="center")

THIN   = Side(style="thin", color="C9C9C9")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

POSITION_COLORS = {
    "QB":  "D6E4F0", "RB":  "D5F5E3", "WR":  "FEF9E7",
    "TE":  "F9EBEA", "K":   "F2F3F4", "DEF": "EAF2FF",
    "DL":  "F5EEF8", "LB":  "F5EEF8", "DB":  "F5EEF8",
}

ROUND_COLORS = {
    1: "FFD700",   # gold   – 1st round
    2: "C0C0C0",   # silver – 2nd round
    3: "CD7F32",   # bronze – 3rd round
    4: "E8F4FD",   # light blue – 4th
    5: "F5F5F5",   # grey   – 5th
}

def pos_fill(pos):
    return PatternFill("solid", start_color=POSITION_COLORS.get(pos, "FFFFFF"))

def round_fill(rnd):
    return PatternFill("solid", start_color=ROUND_COLORS.get(rnd, "FFFFFF"))


# ── API helpers ───────────────────────────────────────────────────────────────

def get(url):
    r = requests.get(url, timeout=30)
    r.raise_for_status()
    return r.json()

def fetch_players():
    if os.path.exists(PLAYERS_CACHE):
        age_h = (datetime.now().timestamp() - os.path.getmtime(PLAYERS_CACHE)) / 3600
        if age_h < 24:
            print(f"  Using cached player data ({age_h:.1f}h old)")
            with open(PLAYERS_CACHE) as f:
                return json.load(f)
    print("  Downloading full player database (~20MB, cached for 24h)...")
    data = get(f"{BASE}/players/{SPORT}")
    with open(PLAYERS_CACHE, "w") as f:
        json.dump(data, f)
    return data

def fetch_season_stats():
    """Fetch full-season stats for STATS_SEASON. Cached for 24h."""
    if os.path.exists(STATS_CACHE):
        age_h = (datetime.now().timestamp() - os.path.getmtime(STATS_CACHE)) / 3600
        if age_h < 24:
            print(f"  Using cached {STATS_SEASON} stats ({age_h:.1f}h old)")
            with open(STATS_CACHE) as f:
                return json.load(f)
    print(f"  Downloading {STATS_SEASON} season stats...")
    data = get(f"{BASE}/stats/nfl/regular/{STATS_SEASON}")
    with open(STATS_CACHE, "w") as f:
        json.dump(data, f)
    return data

def fetch_fantasycalc():
    """Fetch dynasty AND redraft values from FantasyCalc (Superflex, half-PPR, 12 teams).
    Returns:
      fc_values  : {sleeper_id: {value, overallRank, positionRank, trend30Day, tier,
                                 redraftValue, redraftRank, redraftPositionRank}}
      fc_rookies : list of rookie prospect entries (maybeYoe == 0, position != PICK)
      fc_picks   : {pick_name: value}  e.g. {"2026 Pick 1.05": 3519}
    """
    FC_DYN_CACHE    = KTC_CACHE.replace(".json", "_raw.json")
    FC_REDRAFT_CACHE = KTC_CACHE.replace(".json", "_redraft_raw.json")

    # ── Dynasty data ──────────────────────────────────────────────────────────
    raw = None
    if os.path.exists(FC_DYN_CACHE):
        age_h = (datetime.now().timestamp() - os.path.getmtime(FC_DYN_CACHE)) / 3600
        if age_h < 12:
            print(f"  Using cached FC dynasty values ({age_h:.1f}h old)")
            with open(FC_DYN_CACHE) as f:
                raw = json.load(f)

    if raw is None:
        print("  Downloading FantasyCalc dynasty values (Superflex, half-PPR, 12 teams)...")
        raw = get("https://api.fantasycalc.com/values/current?isDynasty=true&numQbs=2&ppr=0.5&numTeams=12")
        with open(FC_DYN_CACHE, "w") as f:
            json.dump(raw, f)

    # ── Redraft data ──────────────────────────────────────────────────────────
    raw_redraft = None
    if os.path.exists(FC_REDRAFT_CACHE):
        age_h = (datetime.now().timestamp() - os.path.getmtime(FC_REDRAFT_CACHE)) / 3600
        if age_h < 12:
            print(f"  Using cached FC redraft values ({age_h:.1f}h old)")
            with open(FC_REDRAFT_CACHE) as f:
                raw_redraft = json.load(f)

    if raw_redraft is None:
        print("  Downloading FantasyCalc redraft values (Superflex, half-PPR, 12 teams)...")
        try:
            raw_redraft = get("https://api.fantasycalc.com/values/current?isDynasty=false&numQbs=2&ppr=0.5&numTeams=12")
            with open(FC_REDRAFT_CACHE, "w") as f:
                json.dump(raw_redraft, f)
        except Exception as e:
            print(f"  [WARNING] FC redraft fetch failed: {e}")
            raw_redraft = []

    # Build redraft lookup by sleeperId
    redraft_by_sid = {}
    for entry in (raw_redraft or []):
        pl  = entry.get("player") or {}
        sid = pl.get("sleeperId")
        if sid:
            redraft_by_sid[sid] = {
                "redraftValue":        entry.get("value"),
                "redraftRank":         entry.get("overallRank"),
                "redraftPositionRank": entry.get("positionRank"),
            }

    # ── Build fc_values, merging dynasty + redraft ────────────────────────────
    fc_values  = {}
    fc_rookies = []
    fc_picks   = {}

    for entry in raw:
        pl  = entry.get("player") or {}
        pos = pl.get("position", "")
        sid = pl.get("sleeperId")

        if pos == "PICK":
            fc_picks[pl.get("name", "")] = entry.get("value", 0)
            continue

        if sid:
            fc_values[sid] = {
                "value":               entry.get("value"),
                "overallRank":         entry.get("overallRank"),
                "positionRank":        entry.get("positionRank"),
                "trend30Day":          entry.get("trend30Day"),
                "tier":                entry.get("maybeTier"),
                "maybeTradeFrequency": entry.get("maybeTradeFrequency"),
                # Redraft fields — None if player not in redraft rankings
                "redraftValue":        redraft_by_sid.get(sid, {}).get("redraftValue"),
                "redraftRank":         redraft_by_sid.get(sid, {}).get("redraftRank"),
                "redraftPositionRank": redraft_by_sid.get(sid, {}).get("redraftPositionRank"),
            }

        if pl.get("maybeYoe") == 0:
            fc_rookies.append({
                "name":        pl.get("name", ""),
                "position":    pos,
                "team":        pl.get("maybeTeam") or "Prospect",
                "age":         pl.get("maybeAge"),
                "sleeperId":   sid,
                "value":       entry.get("value"),
                "overallRank": entry.get("overallRank"),
                "positionRank":entry.get("positionRank"),
                "trend30Day":  entry.get("trend30Day"),
                "tier":        entry.get("maybeTier"),
            })

    fc_rookies.sort(key=lambda r: r.get("overallRank") or 9999)
    print(f"  Loaded {len(fc_values):,} dynasty + {len(redraft_by_sid):,} redraft values, "
          f"{len(fc_rookies)} rookies, {len(fc_picks)} pick values from FantasyCalc")
    return fc_values, fc_rookies, fc_picks

def compute_fantasy_pts(raw_stats, scoring):
    """Dot-product of raw stat keys against league scoring weights."""
    if not raw_stats:
        return None
    total = sum(raw_stats.get(k, 0) * w for k, w in scoring.items())
    return round(total, 1) if total else None

def build_pos_ranks(all_player_pts):
    """Return {player_id: pos_rank} ranked within position by pts descending."""
    by_pos = {}
    for pid, (pos, pts) in all_player_pts.items():
        if pts is not None:
            by_pos.setdefault(pos, []).append((pts, pid))
    ranks = {}
    for pos, entries in by_pos.items():
        entries.sort(reverse=True)
        for rank, (_, pid) in enumerate(entries, 1):
            ranks[pid] = rank
    return ranks

def fetch_all_traded_picks(league_id):
    """
    Resolve current pick ownership across the full league history.

    Sleeper traded_picks field semantics (confirmed empirically):
      roster_id         = ORIGINAL team whose draft slot this pick represents
                          (the unique identifier for the pick)
      owner_id          = CURRENT holder of the pick
      previous_owner_id = who held it immediately before the current holder

    Each (season, round, roster_id) is ONE unique pick. Multiple entries for
    the same pick appear when it has been traded more than once — each entry
    is a historical snapshot of a past owner. The true current holder is the
    owner_id in the entry whose owner_id does NOT appear as previous_owner_id
    in any other entry for that same pick. For single-entry picks the sole
    owner_id is always the current holder.
    """
    # Walk the league chain (current → prev → prev-prev) collecting all picks
    all_picks = []
    lid = league_id
    depth = 0
    while lid and lid != "0" and depth < 4:
        picks = get(f"{BASE}/league/{lid}/traded_picks")
        all_picks.append((depth, picks))
        lid = get(f"{BASE}/league/{lid}").get("previous_league_id")
        depth += 1

    # Group by (season, round, roster_id) — roster_id identifies the pick
    groups = {}
    for depth, picks in all_picks:
        for tp in picks:
            key = (tp["season"], tp["round"], tp["roster_id"])
            groups.setdefault(key, []).append((depth, tp))

    # Resolve current holder per pick
    ownership = {}
    for key, entries in groups.items():
        entries.sort(key=lambda x: x[0])   # current league first
        final = [e for _, e in entries]

        # Build set of owner_ids that appear as previous_owner_id in another entry
        # — those teams have since traded the pick away, so they are NOT current
        prev_owners = {e["previous_owner_id"] for e in final}
        candidates = [e["owner_id"] for e in final if e["owner_id"] not in prev_owners]

        if candidates:
            # Take the candidate from the most-current league entry
            for _, e in entries:
                if e["owner_id"] in candidates:
                    ownership[key] = e["owner_id"]
                    break
        else:
            # Fallback: circular or ambiguous — use most-current league entry
            ownership[key] = entries[0][1]["owner_id"]

    return ownership

def fetch_league_data(league_id):
    print("  Fetching league info...")
    league  = get(f"{BASE}/league/{league_id}")
    print("  Fetching rosters...")
    rosters = get(f"{BASE}/league/{league_id}/rosters")
    print("  Fetching users...")
    users   = get(f"{BASE}/league/{league_id}/users")
    print("  Fetching traded picks (resolving full history)...")
    traded_ownership = fetch_all_traded_picks(league_id)
    print("  Fetching draft settings...")
    drafts  = get(f"{BASE}/league/{league_id}/drafts")

    # Build slot_map: {season -> {roster_id -> draft_slot}} for drafts with known order
    slot_map = {}
    for d in drafts:
        season = d.get("season")
        draft_order = d.get("draft_order") or {}   # user_id -> slot
        if not draft_order:
            continue
        # Map user_id -> roster_id
        uid_to_rid = {r.get("owner_id"): r["roster_id"] for r in rosters}
        slot_map[season] = {uid_to_rid[uid]: slot for uid, slot in draft_order.items() if uid in uid_to_rid}

    return league, rosters, users, traded_ownership, drafts, slot_map


# ── Excel helpers ─────────────────────────────────────────────────────────────

def write_header_row(ws, row, headers, col_start=1):
    for i, h in enumerate(headers):
        c = ws.cell(row=row, column=col_start + i, value=h)
        c.font = HEADER_FONT
        c.fill = HEADER_FILL
        c.alignment = CENTER
        c.border = BORDER

def write_cell(ws, row, col, value, fill=None, bold=False, align=None):
    c = ws.cell(row=row, column=col, value=value)
    c.font      = BOLD_FONT if bold else BODY_FONT
    c.fill      = fill or WHITE_FILL
    c.alignment = align or LEFT
    c.border    = BORDER
    return c

def set_col_widths(ws, widths):
    for col, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(col)].width = w

def add_title_row(ws, title, ncols):
    ws.insert_rows(1)
    c = ws.cell(row=1, column=1, value=title)
    c.font      = Font(name="Arial", bold=True, size=12, color="1F3864")
    c.alignment = CENTER
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=ncols)


# ── Sheet 1: Rosters ──────────────────────────────────────────────────────────

ROSTER_HEADERS = [
    "Team Name", "Owner", "Slot", "Player Name",
    "Position", "NFL Team", "Age", "Experience", "Status",
    f"{STATS_SEASON} Pts", "Pos Rank",
    "FC Value", "FC Rank", "30d Trend", "Tier"
]

def build_rosters_sheet(ws, rosters, users, players, player_pts, pos_ranks, fc_values):
    ws.title = "Rosters"
    write_header_row(ws, 1, ROSTER_HEADERS)

    user_map = {u["user_id"]: u for u in users}

    TREND_UP   = Font(name="Arial", size=10, color="1E8449")   # green  – rising
    TREND_DOWN = Font(name="Arial", size=10, color="C0392B")   # red    – falling

    row = 2
    for roster in sorted(rosters, key=lambda r: r["roster_id"]):
        owner_id   = roster.get("owner_id") or ""
        user       = user_map.get(owner_id, {})
        team_name  = (user.get("metadata") or {}).get("team_name") or user.get("display_name") or f"Team {roster['roster_id']}"
        owner_name = user.get("display_name") or "—"

        player_ids = roster.get("players") or []
        starters   = set(roster.get("starters") or [])
        taxi       = set(roster.get("taxi") or [])

        if not player_ids:
            continue

        sort_key = lambda p: (players.get(p, {}).get("position", "ZZ"), players.get(p, {}).get("last_name", ""))
        for pid in sorted(player_ids, key=sort_key):
            p        = players.get(pid, {})
            name     = f"{p.get('first_name', '')} {p.get('last_name', '')}".strip() or pid
            pos      = p.get("position") or "—"
            age      = p.get("age") or "—"
            exp      = p.get("years_exp")
            exp      = "Rookie" if exp == 0 else (str(exp) + "yr" if exp else "—")
            nfl_team = p.get("team") or "FA"
            status   = p.get("injury_status") or p.get("status") or "Active"
            slot     = "Starter" if pid in starters else ("Taxi" if pid in taxi else "Bench")
            fill     = pos_fill(pos)
            pts      = player_pts.get(pid)
            pts_disp  = pts if pts is not None else "—"
            rank_disp = pos_ranks.get(pid) or "—"

            fc       = fc_values.get(pid, {})
            fc_val   = fc.get("value")   or "—"
            fc_rank  = fc.get("overallRank") or "—"
            fc_tier  = fc.get("tier")    or "—"
            trend    = fc.get("trend30Day")

            for col, val in enumerate([team_name, owner_name, slot, name, pos, nfl_team, age, exp, status], 1):
                write_cell(ws, row, col, val, fill=fill)
            write_cell(ws, row, 10, pts_disp,  fill=fill, align=CENTER)
            write_cell(ws, row, 11, rank_disp, fill=fill, align=CENTER)
            write_cell(ws, row, 12, fc_val,    fill=fill, align=CENTER)
            write_cell(ws, row, 13, fc_rank,   fill=fill, align=CENTER)
            write_cell(ws, row, 15, fc_tier,   fill=fill, align=CENTER)

            # Trend cell with directional colour
            trend_disp = f"+{trend}" if trend and trend > 0 else (str(trend) if trend is not None else "—")
            tc = write_cell(ws, row, 14, trend_disp, fill=fill, align=CENTER)
            if trend and trend > 0:
                tc.font = TREND_UP
            elif trend and trend < 0:
                tc.font = TREND_DOWN

            row += 1

    set_col_widths(ws, [22, 18, 9, 24, 9, 9, 6, 11, 12, 11, 9, 9, 9, 10, 6])
    add_title_row(ws, f"Dynasty Rosters  •  Kingston Dynasty League  •  {datetime.now().strftime('%Y-%m-%d')}", len(ROSTER_HEADERS))
    ws.auto_filter.ref = f"A2:{get_column_letter(len(ROSTER_HEADERS))}2"
    ws.freeze_panes = "A3"


# ── Sheet 2: Draft Picks ──────────────────────────────────────────────────────

PICKS_HEADERS = [
    "Team Name", "Owner", "Season", "Round", "Pick Label",
    "Original Team", "Acquired Via Trade", "FC Value"
]

def build_picks_sheet(ws, rosters, users, traded_ownership, drafts, slot_map, fc_picks):
    ws.title = "Draft Picks"

    # Determine number of rounds from the most recent draft (default 5)
    num_rounds = 5
    if drafts:
        num_rounds = drafts[0].get("settings", {}).get("rounds", 5)

    # Future seasons to display
    current_year = datetime.now().year
    seasons = [str(current_year + i) for i in range(PICK_SEASONS)]

    # Build ownership map: start with each team owning their own picks.
    # Key: (season, round, original_roster_id)  Value: current_holder_roster_id
    ownership = {}
    for r in rosters:
        rid = r["roster_id"]
        for season in seasons:
            for rnd in range(1, num_rounds + 1):
                ownership[(season, rnd, rid)] = rid

    # Apply resolved trades
    for (season, rnd, orig_rid), current_rid in traded_ownership.items():
        key = (season, rnd, orig_rid)
        if key in ownership:
            ownership[key] = current_rid

    # Map roster_id -> team/owner info
    user_map    = {u["user_id"]: u for u in users}
    roster_info = {}
    for r in rosters:
        owner_id   = r.get("owner_id") or ""
        user       = user_map.get(owner_id, {})
        team_name  = (user.get("metadata") or {}).get("team_name") or user.get("display_name") or f"Team {r['roster_id']}"
        owner_name = user.get("display_name") or "—"
        roster_info[r["roster_id"]] = (team_name, owner_name)

    # Collect rows
    pick_rows = []
    for (season, rnd, orig_rid), current_rid in ownership.items():
        orig_team             = roster_info.get(orig_rid, (f"Team {orig_rid}", "—"))[0]
        curr_team, curr_owner = roster_info.get(current_rid, (f"Team {current_rid}", "—"))
        acquired              = "Yes" if orig_rid != current_rid else "No"

        # Use exact slot notation (e.g. 1.05) if draft order is known for this season
        season_slots = slot_map.get(season, {})
        if season_slots and orig_rid in season_slots:
            slot = season_slots[orig_rid]
            pick_label = f"{rnd}.{slot:02d}"
            # Match FantasyCalc pick name e.g. "2026 Pick 1.05"
            fc_pick_key = f"{season} Pick {rnd}.{slot:02d}"
        else:
            pick_label  = f"{season} Round {rnd}"
            fc_pick_key = None

        fc_val = fc_picks.get(fc_pick_key) if fc_pick_key else None

        pick_rows.append((curr_team, curr_owner, season, rnd, pick_label, orig_team, acquired, fc_val, current_rid))

    # Sort by team, season, then pick slot (numeric sort on label)
    def sort_key(r):
        label = r[4]
        if "." in label:
            rnd_part, slot_part = label.split(".")
            return (r[0], r[2], int(rnd_part), int(slot_part))
        return (r[0], r[2], r[3], 99)

    pick_rows.sort(key=sort_key)

    write_header_row(ws, 1, PICKS_HEADERS)

    for i, (team, owner, season, rnd, label, orig_team, acquired, fc_val, _) in enumerate(pick_rows):
        row      = i + 2
        fill     = round_fill(rnd)
        acq_fill = PatternFill("solid", start_color="D5F5E3") if acquired == "Yes" else fill
        for col, val in enumerate([team, owner, season, rnd, label, orig_team], 1):
            write_cell(ws, row, col, val, fill=fill, align=CENTER if col in (3, 4) else LEFT)
        write_cell(ws, row, 7, acquired,            fill=acq_fill, align=CENTER)
        write_cell(ws, row, 8, fc_val if fc_val else "—", fill=fill, align=CENTER)

    set_col_widths(ws, [22, 18, 8, 7, 16, 22, 16, 10])
    add_title_row(ws, f"Draft Picks  •  Seasons {seasons[0]}–{seasons[-1]}  •  {datetime.now().strftime('%Y-%m-%d')}", len(PICKS_HEADERS))
    ws.auto_filter.ref = f"A2:{get_column_letter(len(PICKS_HEADERS))}2"
    ws.freeze_panes = "A3"

    return len(pick_rows)


# ── Sheet 3: Free Agents / Waivers ────────────────────────────────────────────

FA_HEADERS = [
    "Player Name", "Position", "NFL Team", "Age", "Experience", "Status",
    f"{STATS_SEASON} Pts", "Pos Rank", "FC Value", "FC Rank", "30d Trend", "Tier", "Injury Notes"
]

DYNASTY_POSITIONS = {"QB", "RB", "WR", "TE", "K", "DEF", "DL", "LB", "DB"}

def build_fa_sheet(ws, rosters, players, player_pts, pos_ranks, fc_values):
    ws.title = "Free Agents"

    TREND_UP   = Font(name="Arial", size=10, color="1E8449")
    TREND_DOWN = Font(name="Arial", size=10, color="C0392B")

    rostered = set()
    for r in rosters:
        rostered.update(r.get("players") or [])
        rostered.update(r.get("taxi") or [])

    fa_rows = []
    for pid, p in players.items():
        if pid in rostered:
            continue
        pos = p.get("position") or ""
        if pos not in DYNASTY_POSITIONS:
            continue
        if p.get("active") is False and not p.get("team"):
            continue

        name     = f"{p.get('first_name', '')} {p.get('last_name', '')}".strip() or pid
        nfl_team = p.get("team") or "FA/UFA"
        age      = p.get("age") or "—"
        exp      = p.get("years_exp")
        exp      = "Rookie" if exp == 0 else (str(exp) + "yr" if exp else "—")
        status   = p.get("injury_status") or p.get("status") or "Active"
        inj_note = p.get("injury_notes") or "—"
        pts      = player_pts.get(pid)
        rank     = pos_ranks.get(pid)
        fc       = fc_values.get(pid, {})
        fc_val   = fc.get("value") or 0
        fa_rows.append((pos, fc_val, name, pos, nfl_team, age, exp, status, pts, rank, fc, inj_note, pid))

    # Sort by position then FC value descending (best available first)
    pos_order = {"QB": 0, "RB": 1, "WR": 2, "TE": 3, "K": 4, "DEF": 5, "DL": 6, "LB": 7, "DB": 8}
    fa_rows.sort(key=lambda r: (pos_order.get(r[0], 99), -(r[1] or 0)))

    write_header_row(ws, 1, FA_HEADERS)

    for i, (_, _, name, pos, nfl_team, age, exp, status, pts, rank, fc, inj_note, _pid) in enumerate(fa_rows):
        row       = i + 2
        fill      = pos_fill(pos)
        pts_disp  = pts if pts is not None else "—"
        rank_disp = rank if rank is not None else "—"
        fc_val    = fc.get("value")       or "—"
        fc_rank   = fc.get("overallRank") or "—"
        fc_tier   = fc.get("tier")        or "—"
        trend     = fc.get("trend30Day")
        trend_disp = f"+{trend}" if trend and trend > 0 else (str(trend) if trend is not None else "—")

        for col, val in enumerate([name, pos, nfl_team, age, exp, status], 1):
            write_cell(ws, row, col, val, fill=fill)
        write_cell(ws, row, 7,  pts_disp,   fill=fill, align=CENTER)
        write_cell(ws, row, 8,  rank_disp,  fill=fill, align=CENTER)
        write_cell(ws, row, 9,  fc_val,     fill=fill, align=CENTER)
        write_cell(ws, row, 10, fc_rank,    fill=fill, align=CENTER)
        tc = write_cell(ws, row, 11, trend_disp, fill=fill, align=CENTER)
        if trend and trend > 0:
            tc.font = TREND_UP
        elif trend and trend < 0:
            tc.font = TREND_DOWN
        write_cell(ws, row, 12, fc_tier,   fill=fill, align=CENTER)
        write_cell(ws, row, 13, inj_note,  fill=fill)

    set_col_widths(ws, [24, 9, 9, 6, 11, 12, 11, 9, 9, 9, 10, 6, 30])
    add_title_row(ws, f"Free Agents / Waivers  •  {datetime.now().strftime('%Y-%m-%d')}", len(FA_HEADERS))
    ws.auto_filter.ref = f"A2:{get_column_letter(len(FA_HEADERS))}2"
    ws.freeze_panes = "A3"

    return len(fa_rows)


# ── Sheet 4: Scoring Rules ───────────────────────────────────────────────────

# Human-readable labels for every scoring key Sleeper uses
SCORING_LABELS = {
    # Passing
    "pass_yd":       ("Passing",  "Passing Yards (per yd)"),
    "pass_td":       ("Passing",  "Passing TD"),
    "pass_int":      ("Passing",  "Interception Thrown"),
    "pass_2pt":      ("Passing",  "2-Point Conversion (Pass)"),
    "pass_inc":      ("Passing",  "Incomplete Pass"),
    "pass_cmp":      ("Passing",  "Completion"),
    "pass_att":      ("Passing",  "Pass Attempt"),
    "pass_sack":     ("Passing",  "Sacked"),
    "pass_cmp_40p":  ("Passing",  "Completion 40+ Yards"),
    "pass_td_40p":   ("Passing",  "Passing TD 40+ Yards"),
    "pass_td_50p":   ("Passing",  "Passing TD 50+ Yards"),
    # Rushing
    "rush_yd":       ("Rushing",  "Rushing Yards (per yd)"),
    "rush_td":       ("Rushing",  "Rushing TD"),
    "rush_2pt":      ("Rushing",  "2-Point Conversion (Rush)"),
    "rush_att":      ("Rushing",  "Rush Attempt"),
    "rush_fd":       ("Rushing",  "First Down (Rush)"),
    "rush_40p":      ("Rushing",  "Rush 40+ Yards"),
    "rush_td_40p":   ("Rushing",  "Rushing TD 40+ Yards"),
    # Receiving
    "rec":           ("Receiving","Reception"),
    "rec_yd":        ("Receiving","Receiving Yards (per yd)"),
    "rec_td":        ("Receiving","Receiving TD"),
    "rec_2pt":       ("Receiving","2-Point Conversion (Rec)"),
    "rec_fd":        ("Receiving","First Down (Rec)"),
    "rec_40p":       ("Receiving","Reception 40+ Yards"),
    "rec_td_40p":    ("Receiving","Receiving TD 40+ Yards"),
    "bonus_rec_te":  ("Receiving","TE Reception Bonus"),
    "bonus_rec_rb":  ("Receiving","RB Reception Bonus"),
    "bonus_rec_wr":  ("Receiving","WR Reception Bonus"),
    # Misc Offense
    "fum_lost":      ("Misc Off", "Fumble Lost"),
    "fum_rec":       ("Misc Off", "Fumble Recovered"),
    "fum_rec_td":    ("Misc Off", "Fumble Recovery TD"),
    "fum":           ("Misc Off", "Fumble"),
    "st_td":         ("Misc Off", "Special Teams TD"),
    "st_ff":         ("Misc Off", "Special Teams Forced Fumble"),
    "st_fum_rec":    ("Misc Off", "Special Teams Fumble Recovery"),
    "st_tkl_solo":   ("Misc Off", "Special Teams Solo Tackle"),
    # Kicking
    "fgm":           ("Kicking",  "Field Goal Made"),
    "fgmiss":        ("Kicking",  "Field Goal Missed"),
    "fgm_0_19":      ("Kicking",  "FG Made 0-19 Yards"),
    "fgm_20_29":     ("Kicking",  "FG Made 20-29 Yards"),
    "fgm_30_39":     ("Kicking",  "FG Made 30-39 Yards"),
    "fgm_40_49":     ("Kicking",  "FG Made 40-49 Yards"),
    "fgm_50p":       ("Kicking",  "FG Made 50+ Yards"),
    "fgm_50_59":     ("Kicking",  "FG Made 50-59 Yards"),
    "fgm_60p":       ("Kicking",  "FG Made 60+ Yards"),
    "fgmiss_0_19":   ("Kicking",  "FG Missed 0-19 Yards"),
    "fgmiss_50p":    ("Kicking",  "FG Missed 50+ Yards"),
    "xpm":           ("Kicking",  "Extra Point Made"),
    "xpmiss":        ("Kicking",  "Extra Point Missed"),
    # Team Defense
    "pts_allow_0":   ("Team DEF", "Points Allowed: 0"),
    "pts_allow_1_6": ("Team DEF", "Points Allowed: 1-6"),
    "pts_allow_7_13":("Team DEF", "Points Allowed: 7-13"),
    "pts_allow_14_20":("Team DEF","Points Allowed: 14-20"),
    "pts_allow_21_27":("Team DEF","Points Allowed: 21-27"),
    "pts_allow_28_34":("Team DEF","Points Allowed: 28-34"),
    "pts_allow_35p": ("Team DEF", "Points Allowed: 35+"),
    "def_td":        ("Team DEF", "Defensive TD"),
    "int":           ("Team DEF", "Interception"),
    "sack":          ("Team DEF", "Sack"),
    "safe":          ("Team DEF", "Safety"),
    "blk_kick":      ("Team DEF", "Blocked Kick"),
    "ff":            ("Team DEF", "Forced Fumble"),
    "fum_rec":       ("Team DEF", "Fumble Recovery"),
    # IDP
    "idp_tkl_solo":  ("IDP",      "Solo Tackle"),
    "idp_tkl_ast":   ("IDP",      "Assisted Tackle"),
    "idp_tkl_loss":  ("IDP",      "Tackle for Loss"),
    "idp_sack":      ("IDP",      "Sack"),
    "idp_int":       ("IDP",      "Interception"),
    "idp_pass_def":  ("IDP",      "Pass Deflection"),
    "idp_ff":        ("IDP",      "Forced Fumble"),
    "idp_fum_rec":   ("IDP",      "Fumble Recovery"),
    "idp_def_td":    ("IDP",      "Defensive TD"),
    "idp_safe":      ("IDP",      "Safety"),
    "idp_blk_kick":  ("IDP",      "Blocked Kick"),
    "idp_qb_hit":    ("IDP",      "QB Hit"),
}

CATEGORY_COLORS = {
    "Passing":   "DEEAF1",
    "Rushing":   "E2EFDA",
    "Receiving": "FFF2CC",
    "Misc Off":  "FCE4D6",
    "Kicking":   "F2F2F2",
    "Team DEF":  "E9D7F0",
    "IDP":       "D9EAD3",
}

SCORING_HEADERS = ["Category", "Stat", "Points"]

def build_scoring_sheet(ws, scoring):
    ws.title = "Scoring Rules"

    # Only show keys that have a non-zero value
    rows = []
    for key, pts in sorted(scoring.items(), key=lambda x: x[0]):
        if not pts:
            continue
        cat, label = SCORING_LABELS.get(key, ("Other", key))
        rows.append((cat, label, pts))

    # Sort by category order then label
    cat_order = ["Passing", "Rushing", "Receiving", "Misc Off", "Kicking", "Team DEF", "IDP", "Other"]
    rows.sort(key=lambda r: (cat_order.index(r[0]) if r[0] in cat_order else 99, r[1]))

    write_header_row(ws, 1, SCORING_HEADERS)

    for i, (cat, label, pts) in enumerate(rows):
        row  = i + 2
        color = CATEGORY_COLORS.get(cat, "FFFFFF")
        fill  = PatternFill("solid", start_color=color)
        write_cell(ws, row, 1, cat,   fill=fill, align=CENTER)
        write_cell(ws, row, 2, label, fill=fill)
        c = write_cell(ws, row, 3, pts, fill=fill, align=CENTER)
        # Color negative values red
        if pts < 0:
            c.font = Font(name="Arial", size=10, color="FF0000")

    set_col_widths(ws, [14, 32, 10])
    add_title_row(ws, f"League Scoring Rules  •  Kingston Dynasty League", len(SCORING_HEADERS))
    ws.auto_filter.ref = f"A2:{get_column_letter(len(SCORING_HEADERS))}2"
    ws.freeze_panes = "A3"

    return len(rows)


# ── Sheet 5: Trending Players ────────────────────────────────────────────────

TRENDING_HEADERS = [
    "Trend", "Player Name", "Position", "NFL Team", "Age",
    "FC Value", "FC Rank", "2025 Pts", "# Leagues Added/Dropped", "Available in Your League", "Dynasty Team"
]

TREND_LOOKBACK = 48   # hours

def build_trending_sheet(ws, players, rosters, users, player_pts, pos_ranks, fc_values):
    ws.title = "Trending Players"

    user_map = {u["user_id"]: u for u in users}
    pid_to_team = {}
    for r in rosters:
        oid  = r.get("owner_id") or ""
        u    = user_map.get(oid, {})
        team = (u.get("metadata") or {}).get("team_name") or u.get("display_name") or f"Team {r['roster_id']}"
        for pid in (r.get("players") or []):
            pid_to_team[pid] = team
        for pid in (r.get("taxi") or []):
            pid_to_team[pid] = team

    adds  = get(f"{BASE}/players/{SPORT}/trending/add?lookback_hours={TREND_LOOKBACK}&limit=25")
    drops = get(f"{BASE}/players/{SPORT}/trending/drop?lookback_hours={TREND_LOOKBACK}&limit=25")

    rows = []
    for entry in adds:
        rows.append(("Add", entry["player_id"], entry["count"]))
    for entry in drops:
        rows.append(("Drop", entry["player_id"], entry["count"]))

    write_header_row(ws, 1, TRENDING_HEADERS)

    ADD_FILL   = PatternFill("solid", start_color="D5F5E3")
    DROP_FILL  = PatternFill("solid", start_color="FADBD8")
    AVAIL_FILL = PatternFill("solid", start_color="D5F5E3")
    TAKEN_FILL = PatternFill("solid", start_color="F2F3F4")

    for i, (trend, pid, count) in enumerate(rows):
        row  = i + 2
        fill = ADD_FILL if trend == "Add" else DROP_FILL
        p    = players.get(pid, {})
        name = f"{p.get('first_name', '')} {p.get('last_name', '')}".strip() or pid
        pos  = p.get("position") or "—"
        team = p.get("team") or "FA"
        age  = p.get("age") or "—"
        pts_disp     = player_pts.get(pid) or "—"
        fc           = fc_values.get(pid, {})
        fc_val       = fc.get("value")       or "—"
        fc_rank      = fc.get("overallRank") or "—"
        dynasty_team = pid_to_team.get(pid, "—")
        available    = "No" if pid in pid_to_team else "Yes"
        avail_fill   = AVAIL_FILL if available == "Yes" else TAKEN_FILL

        for col, val in enumerate([trend, name, pos, team, age], 1):
            write_cell(ws, row, col, val, fill=fill, align=CENTER if col == 1 else LEFT)
        write_cell(ws, row, 6,  fc_val,       fill=fill,       align=CENTER)
        write_cell(ws, row, 7,  fc_rank,      fill=fill,       align=CENTER)
        write_cell(ws, row, 8,  pts_disp,     fill=fill,       align=CENTER)
        write_cell(ws, row, 9,  count,        fill=fill,       align=CENTER)
        write_cell(ws, row, 10, available,    fill=avail_fill, align=CENTER)
        write_cell(ws, row, 11, dynasty_team, fill=TAKEN_FILL if dynasty_team != "—" else fill)

    set_col_widths(ws, [8, 24, 9, 10, 6, 9, 9, 11, 22, 22, 24])
    add_title_row(ws, f"Trending Players (Last {TREND_LOOKBACK}h across all Sleeper leagues)  •  {datetime.now().strftime('%Y-%m-%d')}", len(TRENDING_HEADERS))
    ws.auto_filter.ref = f"A2:{get_column_letter(len(TRENDING_HEADERS))}2"
    ws.freeze_panes = "A3"


# ── Sheet 6: 2026 Rookies ────────────────────────────────────────────────────

ROOKIES_HEADERS = [
    "FC Rank", "Player Name", "Position", "NFL Team", "Age",
    "FC Value", "Pos Rank", "30d Trend", "Tier", "On Your League Roster"
]

def build_rookies_sheet(ws, fc_rookies, rosters):
    ws.title = "2026 Rookies"

    rostered = set()
    for r in rosters:
        rostered.update(r.get("players") or [])
        rostered.update(r.get("taxi") or [])

    # Build rostered map: sleeper_id -> team name
    pid_to_team = {}
    for r in rosters:
        for pid in (r.get("players") or []) + (r.get("taxi") or []):
            pid_to_team[pid] = r["roster_id"]

    TREND_UP   = Font(name="Arial", size=10, color="1E8449")
    TREND_DOWN = Font(name="Arial", size=10, color="C0392B")
    ROSTERED_FILL = PatternFill("solid", start_color="D5F5E3")
    AVAIL_FILL    = PatternFill("solid", start_color="FFFFFF")

    write_header_row(ws, 1, ROOKIES_HEADERS)

    for i, r in enumerate(fc_rookies):
        row   = i + 2
        pos   = r.get("position", "—")
        fill  = pos_fill(pos)
        sid   = r.get("sleeperId")
        trend = r.get("trend30Day")
        trend_disp = f"+{trend}" if trend and trend > 0 else (str(trend) if trend is not None else "—")
        on_roster  = "Yes" if sid in rostered else "No"
        rf         = ROSTERED_FILL if on_roster == "Yes" else fill
        age        = round(r["age"], 1) if r.get("age") else "—"

        write_cell(ws, row, 1, r.get("overallRank") or "—", fill=fill, align=CENTER)
        write_cell(ws, row, 2, r.get("name", ""),            fill=fill)
        write_cell(ws, row, 3, pos,                          fill=fill, align=CENTER)
        write_cell(ws, row, 4, r.get("team", "Prospect"),    fill=fill, align=CENTER)
        write_cell(ws, row, 5, age,                          fill=fill, align=CENTER)
        write_cell(ws, row, 6, r.get("value") or "—",        fill=fill, align=CENTER)
        write_cell(ws, row, 7, r.get("positionRank") or "—", fill=fill, align=CENTER)
        tc = write_cell(ws, row, 8, trend_disp,              fill=fill, align=CENTER)
        if trend and trend > 0:   tc.font = TREND_UP
        elif trend and trend < 0: tc.font = TREND_DOWN
        write_cell(ws, row, 9,  r.get("tier") or "—",        fill=fill, align=CENTER)
        write_cell(ws, row, 10, on_roster,                   fill=rf,   align=CENTER)

    set_col_widths(ws, [9, 24, 9, 12, 6, 9, 9, 10, 6, 20])
    add_title_row(ws, f"2026 Rookie Rankings (FantasyCalc Superflex)  •  {datetime.now().strftime('%Y-%m-%d')}", len(ROOKIES_HEADERS))
    ws.auto_filter.ref = f"A2:{get_column_letter(len(ROOKIES_HEADERS))}2"
    ws.freeze_panes = "A3"

    return len(fc_rookies)


# ── Main ──────────────────────────────────────────────────────────────────────

def main():
    league_id = LEAGUE_ID.strip()
    if not league_id:
        print("ERROR: Set your LEAGUE_ID at the top of this script.")
        sys.exit(1)

    print(f"\nSleeper Dynasty Tracker — League {league_id}")
    print("=" * 50)

    print("\n[1/3] Fetching league data...")
    league, rosters, users, traded_ownership, drafts, slot_map = fetch_league_data(league_id)
    league_name = league.get("name", f"League {league_id}")
    print(f"      League: {league_name} ({len(rosters)} teams, {len(traded_ownership)} unique picks resolved)")

    print("\n[2/3] Fetching player database, stats & values...")
    players = fetch_players()
    print(f"      Loaded {len(players):,} players")

    raw_stats                    = fetch_season_stats()
    fc_values, fc_rookies, fc_picks = fetch_fantasycalc()
    scoring                      = league.get("scoring_settings") or {}

    player_pts = {}
    all_pts    = {}
    for pid, p in players.items():
        pos  = p.get("position") or "—"
        pts  = compute_fantasy_pts(raw_stats.get(pid), scoring)
        player_pts[pid] = pts
        all_pts[pid]    = (pos, pts)

    pos_ranks = build_pos_ranks(all_pts)
    scored    = sum(1 for v in player_pts.values() if v is not None)
    print(f"      Computed {STATS_SEASON} pts for {scored:,} players (league scoring)")

    print("\n[3/3] Building Excel workbook...")
    wb = Workbook()

    build_rosters_sheet(wb.active, rosters, users, players, player_pts, pos_ranks, fc_values)
    picks_count   = build_picks_sheet(wb.create_sheet(), rosters, users, traded_ownership, drafts, slot_map, fc_picks)
    fa_count      = build_fa_sheet(wb.create_sheet(), rosters, players, player_pts, pos_ranks, fc_values)
    rookies_count = build_rookies_sheet(wb.create_sheet(), fc_rookies, rosters)
    scoring_count = build_scoring_sheet(wb.create_sheet(), scoring)
    build_trending_sheet(wb.create_sheet(), players, rosters, users, player_pts, pos_ranks, fc_values)

    wb.save(OUTPUT)
    print(f"\nDone! Saved to: {os.path.abspath(OUTPUT)}")
    print(f"  Rosters       : {len(rosters)} teams")
    print(f"  Draft Picks   : {picks_count} picks tracked ({len(fc_picks)} with FC values)")
    print(f"  Free Agents   : {fa_count:,} available players")
    print(f"  2026 Rookies  : {rookies_count} prospects ranked")
    print(f"  Scoring Rules : {scoring_count} active scoring categories")
    print(f"  Trending      : top 25 adds + drops (last {TREND_LOOKBACK}h)")


if __name__ == "__main__":
    main()
